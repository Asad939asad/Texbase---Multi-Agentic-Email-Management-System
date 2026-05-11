import pandas as pd
import sqlite3
import requests
import json
import os
import re
import asyncio
from pathlib import Path
# Import the necessary components from your email pipeline script.
# Ensure "emailspecific" matches the actual filename of your email script (without .py).
import sys
# Add the directory containing email_specific.py to the system path
sys.path.append(os.path.join(os.environ.get('WORKSPACE_ROOT', '.'), 'src_2/ColdEmail/'))

from email_specific import run_batch, write_excel, DEEP_RES_PDF_DIR, BATCH_SIZE, EXCEL_OUTPUT
# Outreach data pipeline — no email validation or website scraping needed

# --- CONFIGURATION ---
LLM_URL = "https://unscotched-devon-interpapillary.ngrok-free.dev/generate"
DB_PATH = os.path.join(os.environ.get('WORKSPACE_ROOT', '.'), 'Database/outreach_data/excel_data.db')

os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)

# --- 1. DATABASE SETUP ---
def init_db():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS outreach_companies (
            id                   INTEGER PRIMARY KEY AUTOINCREMENT,
            company_name         TEXT,
            website              TEXT,
            address              TEXT,
            total_shipments      TEXT,
            top_suppliers        TEXT,
            hs_codes             TEXT,
            company_description  TEXT,
            key_executives       TEXT,
            deep_research_pdf    TEXT,
            email                TEXT,
            status               TEXT DEFAULT 'under_review',
            imported_at          DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    conn.close()

# --- 2. LLM INTERACTION ---
def call_llm(system_prompt, user_query):
    payload = {
        "system_prompt": system_prompt,
        "query": user_query,
        "max_new_tokens": 5000 # Keep it short, we only need a JSON object
    }
    response = requests.post(LLM_URL, json=payload)
    if response.status_code == 200:
        return response.json()["response"]
    raise Exception(f"LLM Error: {response.text}")

# --- 3. DYNAMIC COLUMN MAPPING ---
def get_column_mapping(raw_columns):
    print(f"🧠 Asking LLM to map these columns: {raw_columns}")
    
    system_prompt = (
        "You are a precise data engineering assistant. You map raw spreadsheet columns to a target schema. "
        "Output ONLY a valid JSON object. No explanations, no markdown formatting."
    )
    
    user_query = f"""
    Target Schema exact keys required:
    1. "Company Name"
    2. "Website"
    3. "Address"
    4. "Total Shipments"
    5. "Top Suppliers"
    6. "HS Codes"
    7. "Company Description"
    8. "Key Executives"
    9. "Deep Research PDF"
    10. "Email"

    Raw column names found in the file: {raw_columns}

    Match the exact raw column names from the list above to the target schema.
    CRITICAL RULE: The values in your JSON must be EXACT STRINGS chosen ONLY from the 'Raw column names found in the file' list. DO NOT invent or guess column names.
    If a target concept does not exist in the raw columns, map it to null.

    Format example: {{"Company Name": "Company Name", "Website": "Website", "Email": null}}
    """
    
    json_raw = call_llm(system_prompt, user_query)
    
    # Clean the LLM output (Robust JSON parsing)
    json_raw = re.sub(r"```json|```", "", json_raw).strip()
    match = re.search(r'\{.*\}', json_raw, re.DOTALL)
    if match:
        json_raw = match.group(0)
        
    try:
        mapping = json.loads(json_raw)
        print(f"✅ LLM Mapping decided: {json.dumps(mapping, indent=2)}")
        return mapping
    except Exception as e:
        print(f"❌ Failed to parse LLM mapping JSON. Raw output: {json_raw}")
        return None

# --- 4. MAIN PROCESSING LOGIC ---
def process_excel(excel_path):
    print(f"\n📊 Processing Excel: {excel_path}")
    
    # Step A: Load the file — supports .xlsx, .xls and .numbers
    try:
        ext = Path(excel_path).suffix.lower()
        if ext == '.numbers':
            import numbers_parser
            doc = numbers_parser.Document(str(excel_path))
            sheet = doc.sheets[0]
            table = sheet.tables[0]
            rows = list(table.iter_rows())
            headers = [cell.value for cell in rows[0]]
            data = [[cell.value for cell in row] for row in rows[1:]]
            df = pd.DataFrame(data, columns=headers)
            print(f"  → Loaded .numbers file: {len(df)} rows, {len(headers)} columns")
        else:
            df = pd.read_excel(excel_path)
    except Exception as e:
        print(f"❌ Could not read file: {e}")
        return

    # Convert pandas NaN values to None (so SQLite accepts them)
    df = df.where(pd.notnull(df), None)
    
    raw_columns = df.columns.tolist()
    
    # Step B: Get the intelligent mapping from the LLM
    mapping = get_column_mapping(raw_columns)
    if not mapping:
        return

    # Step C: Extract and save to Database
    conn = sqlite3.connect(DB_PATH)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        # Find the 'Status' column index
        status_col = None
        for col_idx in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col_idx).value == 'Status':
                status_col = col_idx
                break
    except Exception as e:
        print(f"⚠️ Could not load workbook for status updates: {e}")
        wb = None

    cursor = conn.cursor()
    records_saved = 0

    for index, row in df.iterrows():
        # 1. Skip if already marked as Completed in the spreadsheet
        if 'Status' in df.columns and row['Status'] == 'Completed':
            continue

        # Resolve each column via the LLM mapping
        def get(key):
            col = mapping.get(key)
            if col and col in df.columns:
                val = row[col]
                if val is None or (isinstance(val, float) and pd.isna(val)):
                    return None
                s = str(val).strip()
                return s if s and s.lower() != 'nan' else None
            return None

        company_name        = get("Company Name")
        website             = get("Website")
        address             = get("Address")
        total_shipments     = get("Total Shipments")
        top_suppliers       = get("Top Suppliers")
        hs_codes            = get("HS Codes")
        company_description = get("Company Description")
        key_executives      = get("Key Executives")
        deep_research_pdf   = get("Deep Research PDF")
        email               = get("Email") or "not updated"

        # Skip completely empty rows
        if not any([company_name, website, address]):
            continue

        # 2. Skip if company already exists in the database (duplicate check)
        cursor.execute('SELECT id FROM outreach_companies WHERE company_name = ? AND website = ?', (company_name, website))
        if cursor.fetchone():
            print(f"  ⏭️ Skipping (already in DB): {company_name}")
            # Ensure Excel is marked completed if it's already in DB but not marked
            if wb and status_col:
                ws.cell(row=index + 2, column=status_col, value='Completed')
            continue

        try:
            cursor.execute('''
                INSERT INTO outreach_companies
                    (company_name, website, address, total_shipments, top_suppliers,
                     hs_codes, company_description, key_executives, deep_research_pdf, email)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (company_name, website, address, total_shipments, top_suppliers,
                  hs_codes, company_description, key_executives, deep_research_pdf, email))
            
            # Update status in the workbook (index + 2 because pandas is 0-based and Excel is 1-based + header)
            if wb and status_col:
                ws.cell(row=index + 2, column=status_col, value='Completed')
            
            records_saved += 1
            print(f"  ✅ Saved: {company_name} | email: {email}")
        except Exception as e:
            print(f"  ❌ Error saving {company_name}: {e}")

    conn.commit()
    conn.close()

    # Step D: Save the Excel file back with updated status
    if wb:
        try:
            wb.save(excel_path)
            print(f"📊 Excel file updated with 'Completed' status: {excel_path}")
        except Exception as e:
            print(f"❌ Could not save updated Excel: {e}")

    print(f"🎉 Successfully saved {records_saved} records to the database.")

if __name__ == "__main__":
    # Ensure DB is initialized before anything else
    init_db()

    print(f"\n{'═'*65}")
    print(f"  TEXBASE Cold Email Pipeline — Batch of {BATCH_SIZE}")
    print(f"{'═'*65}")
    print(f"  📂 WORKSPACE_ROOT: {os.environ.get('WORKSPACE_ROOT', '.')}")
    print(f"  📂 DB_PATH (Source): {DB_PATH}")

    all_states = asyncio.run(run_batch(BATCH_SIZE))

    # Filter out empty states (queue was empty for that slot)
    valid_states = [s for s in all_states if s.get("company_id")]

    if valid_states:
        write_excel(valid_states, EXCEL_OUTPUT)
        print(f"\n{'═'*65}")
        print("  BATCH SUMMARY")
        print(f"{'═'*65}")
        for s in valid_states:
            print(f"  • {s.get('company_name', '?')} — "
                  f"result: {s.get('result', s.get('error', ''))}")
        print(f"\n  Excel output → {EXCEL_OUTPUT}")
    else:
        print("\n  No new companies processed by the generation pipeline.")

    # =========================================================
    # PHASE 2: IMPORT EXCEL DATA TO DB
    # =========================================================
    # We process the Excel file regardless of Phase 1 results (supports manual edits)
    excel_path_to_process = Path(EXCEL_OUTPUT) 
    
    if os.path.exists(excel_path_to_process):
        print(f"\n  📥 Triggering Excel Import for: {excel_path_to_process}")
        process_excel(excel_path_to_process)
    else:
        print(f"❌ Excel file not found for import: {excel_path_to_process}")