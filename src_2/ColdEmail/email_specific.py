#!/usr/bin/env python3
"""
email_specific.py  —  TEXBASE Cold Email Pipeline (LangGraph Architecture)
===========================================================================
Improvements:
  • Parallel processing of 2 companies simultaneously via asyncio
  • Structured JSON email extraction schema — no regex fishing
  • Deep research contact extraction uses an explicit Gemini parse step
  • Excel output (outreach_results.xlsx) written after every batch
  • Rollback on failure per company (independent)

Flow per company (LangGraph StateGraph):
  fetch_company ──► profile_gemini ──► verify_emails ──► deep_research
                                                               │
                    save_to_db ◄── draft_emails ◄── scrape_importyeti
"""

from __future__ import annotations

import asyncio
import threading
import sqlite3
import json
import os
import re
import sys
import time
import requests
import urllib.parse
from dotenv import load_dotenv
load_dotenv(os.path.join(os.environ.get('WORKSPACE_ROOT', '.'), 'backend/.env'))
import datetime
from typing import TypedDict, Optional

from langgraph.graph import StateGraph, END

from google import genai
from google.genai import types

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

sys.path.append(os.path.join(os.environ.get('WORKSPACE_ROOT', '.'), 'USA_ImportYeti'))
from scrape_importyeti_profiles import scrape_single_link, parse_data

# ─── Paths & Keys ─────────────────────────────────────────────────────────────
DB_PATH          = os.path.join(os.environ.get('WORKSPACE_ROOT', '.'), 'USA_ImportYeti/importyeti_data.db')
TRACKER_DB_PATH  = os.path.join(os.environ.get('WORKSPACE_ROOT', '.'), 'src_2/ColdEmail/outreach_tracker.db')
DEEP_RES_PDF_DIR = os.path.join(os.environ.get('WORKSPACE_ROOT', '.'), 'src_2/ColdEmail/deep_research_reports')
EXCEL_OUTPUT     = os.path.join(os.environ.get('WORKSPACE_ROOT', '.'), 'Database/excelSheet/outreach_results.xlsx')
HUNTER_API_KEY   = ""   # not used — email finding removed
GEMINI_API_KEY   = os.getenv("GEMINI_API_KEY_3")

BATCH_SIZE = 1        # companies to process in parallel

# ─── Arooj Enterprises sender profile ─────────────────────────────────────────
SENDER = {
    "company":   "Arooj Enterprises",
    "est":       "1993",
    "name":      "Asad Irfan",
    "title":     "Senior Marketing Manager",
    "website":   "www.texbase.com",
    "capacity":  "150k units/month",
    "certs":     "ISO 14001, SEDEX, and OEKO-TEX",
    "advantage": "vertical integration and rigorous quality control",
}

# ─── Gemini Clients ────────────────────────────────────────────────────────────
client             = genai.Client(api_key=GEMINI_API_KEY)
google_search_tool = types.Tool(google_search=types.GoogleSearch())

FETCH_LOCK = threading.Lock()


# ══════════════════════════════════════════════════════════════════════════════
#  PIPELINE STATE
# ══════════════════════════════════════════════════════════════════════════════

class PipelineState(TypedDict, total=False):
    company_data:  dict
    company_id:    int
    company_name:  str
    enriched_json: dict   # Gemini profile (no emails)
    website:       str
    executives:    list   # [{name, title}, ...]
    deep_text:     str
    pdf_path:      str
    result:        str
    error:         str


# ══════════════════════════════════════════════════════════════════════════════
#  TRACKER DATABASE HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def init_tracker_db() -> sqlite3.Connection:
    """Open (or create) the single outreach_companies table."""
    os.makedirs(os.path.dirname(TRACKER_DB_PATH), exist_ok=True)
    conn = sqlite3.connect(TRACKER_DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS outreach_companies (
            id                     INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id             INTEGER UNIQUE,
            company_name           TEXT,
            website                TEXT,
            address                TEXT,
            total_shipments        TEXT,
            top_suppliers          TEXT,
            hs_codes               TEXT,
            company_description    TEXT,
            key_executives         TEXT,   -- JSON: [{name, title}, ...]
            deep_research_summary  TEXT,
            deep_research_pdf_path TEXT,
            status                 TEXT DEFAULT 'under_review',
            run_date               TEXT
        );
    """)
    conn.commit()
    return conn


def save_company_to_db(conn, state: dict) -> None:
    """Upsert one row per company into outreach_companies."""
    cd           = state.get("company_data", {})
    company_id   = state.get("company_id")
    company_name = state.get("company_name", "")
    enriched     = cd.get("enriched_data", {})
    profile      = enriched.get("company_profile", {})
    website      = profile.get("website", "")
    description  = profile.get("comprehensive_description", "")
    executives   = json.dumps(state.get("executives", []), ensure_ascii=False)
    address      = cd.get("address", "")
    total_ship   = cd.get("total_shipments", "")
    top_sup      = cd.get("top_suppliers", "")
    hs_codes     = cd.get("scraped_profile", {}).get("hs_codes", "")
    deep_text    = state.get("deep_text", "")
    pdf_path     = state.get("pdf_path", "")

    conn.execute("""
        INSERT INTO outreach_companies
            (company_id, company_name, website, address, total_shipments,
             top_suppliers, hs_codes, company_description, key_executives,
             deep_research_summary, deep_research_pdf_path, status, run_date)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'under_review', ?)
        ON CONFLICT(company_id) DO UPDATE SET
            company_name           = excluded.company_name,
            website                = excluded.website,
            address                = excluded.address,
            total_shipments        = excluded.total_shipments,
            top_suppliers          = excluded.top_suppliers,
            hs_codes               = excluded.hs_codes,
            company_description    = excluded.company_description,
            key_executives         = excluded.key_executives,
            deep_research_summary  = excluded.deep_research_summary,
            deep_research_pdf_path = excluded.deep_research_pdf_path,
            run_date               = excluded.run_date
    """, (
        company_id, company_name, website, address, total_ship,
        top_sup, hs_codes, description, executives,
        deep_text, pdf_path, datetime.datetime.now().isoformat(),
    ))
    conn.commit()



# ══════════════════════════════════════════════════════════════════════════════
#  DEEP RESEARCH
# ══════════════════════════════════════════════════════════════════════════════

def clean_markdown_for_pdf(text: str) -> str:
    text = re.sub(r"#{1,6}\s*", "", text)
    text = re.sub(r"\*{1,3}([^*]+)\*{1,3}", r"\1", text)
    text = re.sub(r"__([^_]+)__", r"\1", text)
    text = re.sub(r"\[([^\]]+)\]\([^)]+\)", r"\1", text)
    text = re.sub(r"`{1,3}[^`]*`{1,3}", "", text)
    text = re.sub(r"^[-*+]\s+", "• ", text, flags=re.MULTILINE)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def _save_deep_research_pdf(company_name: str, research_text: str) -> str:
    os.makedirs(DEEP_RES_PDF_DIR, exist_ok=True)
    safe_name = re.sub(r"[^a-zA-Z0-9_\-]", "_", company_name)[:60]
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    try:
        from fpdf import FPDF  # type: ignore
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        pdf.set_font("Helvetica", style="B", size=16)
        pdf.cell(0, 10, txt=f"Deep Research: {company_name}", ln=True)
        pdf.set_font("Helvetica", size=9)
        pdf.cell(0, 6, txt=f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}", ln=True)
        pdf.ln(4)
        pdf.set_font("Helvetica", size=11)
        clean_text = clean_markdown_for_pdf(research_text)
        safe_text  = clean_text.encode("latin-1", "replace").decode("latin-1")
        pdf.multi_cell(0, 6, txt=safe_text)
        pdf_path = os.path.join(DEEP_RES_PDF_DIR, f"{safe_name}_{timestamp}.pdf")
        pdf.output(pdf_path)
        print(f"  [Deep Research] PDF saved → {pdf_path}")
        return pdf_path
    except ImportError:
        txt_path = os.path.join(DEEP_RES_PDF_DIR, f"{safe_name}_{timestamp}.txt")
        with open(txt_path, "w", encoding="utf-8") as f:
            f.write(f"Deep Research Report: {company_name}\n")
            f.write(f"Generated: {datetime.datetime.now()}\n\n")
            f.write(research_text)
        print(f"  [Deep Research] Text report saved → {txt_path}")
        return txt_path


def run_deep_research(company_data: dict) -> tuple:
    """Run deep research and return (research_text, pdf_path). Uses polling pattern."""
    company_name    = company_data.get("company_name", "Unknown")
    address         = company_data.get("address", "")
    total_shipments = company_data.get("total_shipments", "")
    recent_shipment = company_data.get("recent_shipment", "")
    top_suppliers   = company_data.get("top_suppliers", "")
    profile_url     = company_data.get("profile_url", "")
    website         = company_data.get("enriched_data", {}).get("company_profile", {}).get("website", "")

    print(f"\n[Deep Research] Starting for '{company_name}'…")

    query = f"""You are a corporate intelligence researcher. Using ALL available sources
(company website, LinkedIn, ZoomInfo, Apollo, Crunchbase, news articles, press releases),
research the following US-based import company and compile a detailed intelligence report.

═══ COMPANY INTELLIGENCE BRIEF ═══
Company Name    : {company_name}
Address         : {address or 'See ImportYeti profile'}
Total Shipments : {total_shipments}
Most Recent PO  : {recent_shipment}
Known Suppliers : {top_suppliers}
ImportYeti URL  : {profile_url}
Website         : {website or 'Search to find it'}
══════════════════════════════════

Please research and provide:

1. EXECUTIVE TEAM — Full names, titles, and LinkedIn URLs for:
   - CEO / President / Owner
   - VP / Director of Purchasing or Sourcing
   - Merchandising Manager / Head of Product
   - Any other buying / procurement contacts

2. COMPANY PROFILE — Official website, main product categories imported,
   typical order sizes, primary countries of origin sourced from.

3. SUPPLIER CONTEXT — What do their known suppliers ({top_suppliers}) specialise in?

4. MARKET POSITION — Estimated revenue, key retail partners, brand positioning.

Be thorough and precise. Do NOT include email addresses.
"""

    MAX_POLL_SECONDS = 1200  # 20-minute absolute ceiling

    try:
        # ── Step 1: Create the background interaction ──────────────────────
        interaction = client.interactions.create(
            input=query,
            agent="deep-research-preview-04-2026",
            background=True,
        )
        print(f"  [Deep Research] Interaction started: {interaction.id}")

        # ── Step 2: Poll until completed or failed ─────────────────────────
        start_time = time.time()
        dot_count  = 0
        while True:
            elapsed = time.time() - start_time

            if elapsed > MAX_POLL_SECONDS:
                print(f"  [Deep Research] ⚠️  Max timeout ({MAX_POLL_SECONDS}s) reached.")
                return "", ""

            interaction = client.interactions.get(interaction.id)
            status = interaction.status

            if status == "completed":
                print(f"\n  [Deep Research] ✅ Research complete for '{company_name}'")
                break
            elif status == "failed":
                error = getattr(interaction, "error", "unknown error")
                print(f"  [Deep Research] ❌ Research failed: {error}")
                return "", ""
            else:
                # Still running — print a heartbeat dot every 10s
                dot_count += 1
                print(f"  [Deep Research] ⏳ [{elapsed:.0f}s] Status: {status}…", flush=True)
                time.sleep(10)

        # ── Step 3: Extract text from outputs ─────────────────────────────
        research_text = ""
        outputs = getattr(interaction, "outputs", None) or []
        if outputs:
            research_text = outputs[-1].text or ""

        if not research_text:
            print("  [Deep Research] ⚠️  No text in outputs. Interaction may be empty.")

        pdf_path = _save_deep_research_pdf(company_name, research_text)
        print(f"  [Deep Research] Done — report saved.")
        return research_text, pdf_path

    except Exception as e:
        print(f"  [Deep Research] Error: {e}")
        return "", ""




# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL OUTPUT  (one row per company)
# ══════════════════════════════════════════════════════════════════════════════

EXCEL_COLUMNS = [
    "Run Date",
    "Company ID",
    "Company Name",
    "Website",
    "Address",
    "Total Shipments",
    "Top Suppliers",
    "HS Codes",
    "Company Description",
    "Key Executives",        # JSON: [{name, title}, ...]
    "Deep Research PDF",
    "Status",                # default: under_review
]


def write_excel(all_states: list[dict], output_path: str) -> None:
    """
    Write / append one row per company to the Excel file.
    Deduplicates by company_id — if the company already exists, its row is skipped.
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    if os.path.exists(output_path):
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Outreach"
        _write_header(ws)

    # ── Build dedup set from existing rows (col 2 = Company ID) ────
    existing_ids: set = set()
    for row_idx in range(2, ws.max_row + 1):
        cid = ws.cell(row=row_idx, column=2).value
        if cid is not None:
            existing_ids.add(str(cid).strip())

    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    added_count = 0
    for state in all_states:
        company_id = state.get("company_id")
        if not company_id:
            continue
        
        cid_str = str(company_id).strip()
        if cid_str in existing_ids:
            print(f"  [Excel] Skipping duplicate row for company_id={cid_str} ({state.get('company_name')})")
            continue
        
        existing_ids.add(cid_str)
        added_count += 1

        cd           = state.get("company_data", {})
        company_name = state.get("company_name", "")
        enriched     = cd.get("enriched_data", {})
        profile      = enriched.get("company_profile", {})
        website      = profile.get("website", "")
        description  = profile.get("comprehensive_description", "")
        address      = cd.get("address", "")
        total_ship   = cd.get("total_shipments", "")
        top_sup      = cd.get("top_suppliers", "")
        hs_codes     = cd.get("scraped_profile", {}).get("hs_codes", "")
        executives   = json.dumps(state.get("executives", []), ensure_ascii=False)
        pdf_path     = state.get("pdf_path", "")
        run_date     = datetime.date.today().isoformat()

        row_data = [
            run_date, company_id, company_name, website, address,
            total_ship, top_sup, hs_codes, description, executives,
            pdf_path, "under_review",
        ]
        _append_row(ws, row_data, border)

    wb.save(output_path)
    print(f"\n[Excel] Saved {added_count} new entries → {output_path}")


def _write_header(ws) -> None:
    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    for col_idx, col_name in enumerate(EXCEL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        1: 12,   # Run Date
        2: 10,   # Company ID
        3: 30,   # Company Name
        4: 28,   # Website
        5: 32,   # Address
        6: 16,   # Total Shipments
        7: 36,   # Top Suppliers
        8: 22,   # HS Codes
        9: 60,   # Company Description
        10: 50,  # Key Executives
        11: 50,  # Deep Research PDF
        12: 16,  # Status
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def _append_row(ws, row_data: list, border) -> None:
    row_idx  = ws.max_row + 1
    alt_fill = PatternFill("solid", start_color="EBF3FB") if row_idx % 2 == 0 else None
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font      = Font(name="Arial", size=9)
        cell.border    = border
        cell.alignment = Alignment(vertical="top", wrap_text=(col_idx in (9, 10)))
        if alt_fill:
            cell.fill = alt_fill
    ws.row_dimensions[row_idx].height = 18



# ══════════════════════════════════════════════════════════════════════════════
#  LANGGRAPH NODES
# ══════════════════════════════════════════════════════════════════════════════

def node_fetch_company(state: PipelineState) -> PipelineState:
    print("\n" + "═"*65)
    print("  [Node 1/7] fetch_company")
    print("═"*65)

    with FETCH_LOCK:
        try:
            print(f"  [DEBUG] Attempting to open DB: {os.path.abspath(DB_PATH)}")
            if not os.path.exists(DB_PATH):
                print(f"  [ERROR] Database file does not exist at: {os.path.abspath(DB_PATH)}")
            
            conn = sqlite3.connect(DB_PATH)
            conn.row_factory = sqlite3.Row

            # ── Check tracker DB to avoid processing duplicates ──────
            # We attach the tracker DB to do a single query check
            exclude_ids = []
            if os.path.exists(TRACKER_DB_PATH):
                try:
                    conn.execute("ATTACH DATABASE ? AS tracker", (TRACKER_DB_PATH,))
                    # Verify table exists before querying
                    table_check = conn.execute(
                        "SELECT name FROM tracker.sqlite_master WHERE type='table' AND name='outreach_companies'"
                    ).fetchone()
                    if table_check:
                        exclude_query = "AND id NOT IN (SELECT company_id FROM tracker.outreach_companies)"
                    else:
                        exclude_query = ""
                except Exception as e:
                    print(f"  [WARN] Could not attach tracker DB: {e}")
                    exclude_query = ""
            else:
                exclude_query = ""

            query = f"""
                SELECT * FROM companies
                WHERE (selection = 0 OR selection IS NULL)
                {exclude_query}
                ORDER BY CAST(REPLACE(total_shipments, ',', '') AS INTEGER) DESC
                LIMIT 1
            """

            row = conn.execute(query).fetchone()

            if not row:
                conn.close()
                return {**state, "error": "No companies left in queue.", "result": "Queue empty."}

            company_data = dict(row)
            company_id   = company_data["id"]

            # Mark as selected immediately
            conn.execute("UPDATE companies SET selection = 1 WHERE id = ?", (company_id,))
            conn.commit()
            conn.close()

            print(f"  → Fetched: {company_data['company_name']} (id={company_id})")
            return {
                **state,
                "company_data":  company_data,
                "company_id":    company_id,
                "company_name":  company_data["company_name"],
                "enriched_json": {},
                "executives":    [],
                "website":       "",
                "deep_text":     "",
                "pdf_path":      "",
                "error":         "",
            }
        except Exception as e:
            print(f"  [ERROR] fetch_company node: {e}")
            return {**state, "error": str(e)}


def node_profile_gemini(state: PipelineState) -> PipelineState:
    print("\n" + "─"*65)
    print("  [Node 2/4] profile_gemini")
    print("─"*65)
    company_data     = state["company_data"]
    company_json_str = json.dumps(company_data, indent=2)

    profile_prompt = f"""**ROLE:** Corporate Intelligence Analyst.
**GOAL:** Verify US presence, build a company profile, and identify key executives.

Company data:
```json
{company_json_str}
```

**TASKS:**
1. Verify real US commercial location and address.
2. Write a comprehensive company description (products imported, market segment, key retail partners, annual revenue estimate).
3. Identify the executive team: CEO/President/Owner, VP Purchasing, Merchandising Manager, and any other procurement contacts.
   - For each executive provide: Full Name, Job Title, and LinkedIn URL.
   - Do NOT include or predict any email addresses.

Return ONLY raw JSON (no markdown fences):
{{
  "verification_status": {{"is_usa_based": true, "address_type": "...", "entity_match_confidence": "..."}},
  "company_profile": {{
    "official_name": "...",
    "website": "...",
    "main_phone": "...",
    "comprehensive_description": "..."
  }},
  "executives": [
    {{
      "name": "Full Name",
      "title": "Job Title",
      "linkedin": "https://linkedin.com/in/..."
    }}
  ]
}}"""

    enriched_json = {}
    try:
        resp = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=profile_prompt,
            config=types.GenerateContentConfig(
                tools=[google_search_tool],
                response_modalities=["TEXT"],
            ),
        )
        raw = resp.text.strip()
        raw = re.sub(r"^```(?:json)?\s*", "", raw, flags=re.MULTILINE)
        raw = re.sub(r"```\s*$", "", raw, flags=re.MULTILINE).strip()
        try:
            enriched_json = json.loads(raw)
        except json.JSONDecodeError:
            print("  [WARN] Could not parse Gemini profile JSON.")
    except Exception as e:
        print(f"  [Gemini] Profile error: {e}")

    company_data["enriched_data"] = enriched_json
    website    = enriched_json.get("company_profile", {}).get("website", "")
    executives = enriched_json.get("executives", [])
    print(f"  → {len(executives)} executive(s) identified")
    return {**state, "enriched_json": enriched_json, "company_data": company_data,
            "website": website, "executives": executives}


def node_deep_research(state: PipelineState) -> PipelineState:
    print("\n" + "─"*65)
    print("  [Node 3/4] deep_research")
    print("─"*65)
    company_data = state["company_data"]
    research_text, pdf_path = run_deep_research(company_data)
    return {**state, "deep_text": research_text, "pdf_path": pdf_path}


def node_scrape_importyeti(state: PipelineState) -> PipelineState:
    print("\n" + "─"*65)
    print("  [Node 2b/4] scrape_importyeti")
    print("─"*65)
    company_data = state["company_data"]
    profile_url  = company_data.get("profile_url", "")

    if profile_url and isinstance(profile_url, str) and profile_url.startswith("http"):
        print(f"  Fetching: {profile_url}")
        try:
            html = scrape_single_link(profile_url)
            if html:
                loc, codes = parse_data(html)
                company_data["scraped_profile"] = {"location_info": loc, "hs_codes": codes}
                print(f"  ✓ HS codes: {codes}")
            else:
                company_data["scraped_profile"] = {"error": "No HTML returned"}
        except Exception as e:
            company_data["scraped_profile"] = {"error": str(e)}
    else:
        print("  No profile URL — skipped.")
        company_data["scraped_profile"] = {}

    return {**state, "company_data": company_data}




def node_save_to_db(state: PipelineState) -> PipelineState:
    print("\n" + "─"*65)
    print("  [Node 4/4] save_to_db")
    print("─"*65)
    company_name = state.get("company_name", "?")

    tracker = init_tracker_db()
    save_company_to_db(tracker, state)
    tracker.close()

    summary = f"Done — '{company_name}' | profile saved | status=under_review"
    print(f"\n{'═'*65}\n  {summary}\n{'═'*65}")
    return {**state, "result": summary}


# ══════════════════════════════════════════════════════════════════════════════
#  CONDITIONAL ROUTER
# ══════════════════════════════════════════════════════════════════════════════

def route_after_fetch(state: PipelineState) -> str:
    if state.get("error") or not state.get("company_data"):
        return "end"
    return "continue"


# ══════════════════════════════════════════════════════════════════════════════
#  BUILD LANGGRAPH
# ══════════════════════════════════════════════════════════════════════════════

def build_graph():
    graph = StateGraph(PipelineState)

    graph.add_node("fetch_company",     node_fetch_company)
    graph.add_node("profile_gemini",    node_profile_gemini)
    graph.add_node("scrape_importyeti", node_scrape_importyeti)
    graph.add_node("deep_research",     node_deep_research)
    graph.add_node("save_to_db",        node_save_to_db)

    graph.set_entry_point("fetch_company")

    graph.add_conditional_edges(
        "fetch_company",
        route_after_fetch,
        {"continue": "profile_gemini", "end": END},
    )

    graph.add_edge("profile_gemini",    "scrape_importyeti")
    graph.add_edge("scrape_importyeti", "deep_research")
    graph.add_edge("deep_research",     "save_to_db")
    graph.add_edge("save_to_db",        END)

    return graph.compile()


pipeline = build_graph()


# ══════════════════════════════════════════════════════════════════════════════
#  ROLLBACK HELPER
# ══════════════════════════════════════════════════════════════════════════════

def _rollback_company(company_id: int, reason: str) -> None:
    print(f"\n[ROLLBACK] company_id={company_id} — {reason}")
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.execute("UPDATE companies SET selection = 0 WHERE id = ?", (company_id,))
        conn.commit()
        conn.close()
        print(f"[ROLLBACK] ✓ selection reset to 0 for company_id={company_id}")
    except Exception as rb_err:
        print(f"[ROLLBACK ERROR] {rb_err}")


# ══════════════════════════════════════════════════════════════════════════════
#  ASYNC PARALLEL RUNNER
# ══════════════════════════════════════════════════════════════════════════════

async def run_single_pipeline() -> dict:
    """Run one company through the pipeline in a thread pool (blocking I/O)."""
    loop = asyncio.get_event_loop()
    final_state = {}
    try:
        final_state = await loop.run_in_executor(None, lambda: pipeline.invoke({}))

        error_msg = final_state.get("error", "")
        if error_msg and final_state.get("company_id"):
            _rollback_company(final_state["company_id"], f"error: {error_msg}")

    except Exception as e:
        print(f"\n[CRITICAL ERROR] Pipeline crashed: {e}")
        if final_state and isinstance(final_state, dict) and final_state.get("company_id"):
            _rollback_company(final_state["company_id"], f"crash: {e}")

    return final_state if isinstance(final_state, dict) else {}


async def run_batch(batch_size: int) -> list[dict]:
    """Run `batch_size` companies in parallel and return all final states."""
    tasks = [run_single_pipeline() for _ in range(batch_size)]
    results = await asyncio.gather(*tasks, return_exceptions=False)
    return list(results)


# ══════════════════════════════════════════════════════════════════════════════
#  CLI ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    os.makedirs(DEEP_RES_PDF_DIR, exist_ok=True)

    print(f"\n{'═'*65}")
    print(f"  TEXBASE Cold Email Pipeline — Batch of {BATCH_SIZE}")
    print(f"{'═'*65}")

    all_states = asyncio.run(run_batch(BATCH_SIZE))

    # Filter out empty states (queue was empty for that slot)
    valid_states = [s for s in all_states if s.get("company_id")]

    if valid_states:
        write_excel(valid_states, EXCEL_OUTPUT)
        print(f"\n{'═'*65}")
        print("  BATCH SUMMARY")
        print(f"{'═'*65}")
        for s in valid_states:
            print(f"  • {s.get('company_name', '?')} — "
                  f"result: {s.get('result', s.get('error', ''))}")
        print(f"\n  Excel output → {EXCEL_OUTPUT}")
    else:
        print("\n  No companies processed (queue may be empty).")